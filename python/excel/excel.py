#!/usr/bin/env pthon
#coding=utf-8

# Author: Moro
# blog: https://mrchens.github.io

from openpyxl import load_workbook
import os
import re
import time
import pandas
import sys

def preMonth():
    nowDay = time.localtime()
    month = nowDay[1] - 1 or 12
    preMStr = '%s年%s月' % (time.strftime("%y", nowDay), month)
    return preMStr;

def fillEmptyMonth(data, count, str):
    for i in range(1, count):
        data.append(str)

def sortHeads(heads):
    desHeads = ['周期','投票序','投票人','团队']
    desHeads.extend(heads)
    desHeads.extend(['IP地址', '来源', '开始时间', '结束时间', '浏览器', '操作系统', '答题时长', '状态'])
    return desHeads

def checkCommandIsLegal():
    # .xlsx,.xlsm,.xltx,.xltm openpyxl 支持的格式
    argLen = len(sys.argv)
    if argLen >= 3:
        return False;
    if argLen == 1:
        return False;
    arg = str(sys.argv[1])
    if re.match(r'(.*).csv$', arg):
        return True;
    return False;

def csv2xlsx(csvFile, csvEncoding, xlsxFile, xlsxSheetName):
    csv = pandas.read_csv(csvFile, encoding=csvEncoding)
    csv.to_excel(xlsxFile, sheet_name=xlsxSheetName)

csvFileName = "source.csv"
desSheetName = "destinationSheet"
sourceSheetName = 'sourceSheet'
desExcelName = "destination"
excelExtension = ".xlsx"
destinationExcel = "destination.xlsx"
engineName = "openpyxl"
csvEncode = 'gb18030'

peoples = []
groups = []

# pattern
candidateP = '_开放选项$'
nameP = '——(.*)，'
pureNameP = '——(.*)'
continueP = '.*选择理由$'
rowP = '^Q2_'
groupP = '-(.*)\W\W'

if checkCommandIsLegal() == False:
    print('参数错误，请按照如下格式运行脚本:')
    print('python excel.py xxxx.csv')
    exit(0)

csvFileName = sys.argv[1]

print('\ninput file:' + csvFileName)

if os.path.isfile(destinationExcel):
    os.remove(destinationExcel)

print('\noutput file:' + destinationExcel)

csv2xlsx(csvFileName, csvEncode, destinationExcel, sourceSheetName)
workbook = load_workbook(destinationExcel)

print('\nOpen xlsx file:\n' + os.getcwd() + "/" + destinationExcel)
print('\nHandle sheets:', workbook.sheetnames)
print('\nWorking..........')

orisheet = workbook[workbook.sheetnames[0]]

filterDic = {}
desDic = {}
# FIXME: Moro 内存用太多可能会炸
for i in range(1, orisheet.max_column+1):
    columnName = orisheet.cell(1,i).value;
    if columnName == None:
        continue
    if re.match(continueP, columnName):
        continue
    if re.search(candidateP, columnName):
        findOjb = re.search(nameP, columnName)
        if findOjb:
            columnName = findOjb.group(1)
            peoples.append(columnName)
    if columnName == "答卷编号":
        columnName = "投票序"
    rowFlag = False
    if re.match(rowP, columnName):#投票人
        columnName = "投票人"
        filterDic.setdefault('团队', [])
        rowFlag = True
    filterDic.setdefault(columnName, [])
    for row in range(2, orisheet.max_row+1):
        cellValue = orisheet.cell(row,i).value
        if rowFlag:
            if cellValue.find("-") == -1:
                cellValue = "-" + cellValue
            splits = re.split('-(.*)\W\W',cellValue)
            # findName = re.search(pureNameP, cellValue)
            findGroup = splits[1]
            findName = splits[2]
            if findGroup:
                filterDic['团队'].append(findGroup)
            if findName:
                cellValue = findName
        filterDic[columnName].append(cellValue);

desKeys = sortHeads(peoples)

fillEmptyMonth(filterDic.setdefault('周期',[]), orisheet.max_row, preMonth())

for keys in desKeys:
    desDic.setdefault(keys,filterDic[keys])

filterDic.clear();

df = pandas.DataFrame(desDic)
# print(df.info)
writer = pandas.ExcelWriter(destinationExcel, engine=engineName)
writer.book = workbook
writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)
df.to_excel(writer, sheet_name=desSheetName, index=False, header=True, startrow=0)

writer.save()
writer.close()
workbook.close()
print("\nDone!\n")
exit(0)
